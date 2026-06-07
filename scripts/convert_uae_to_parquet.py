from __future__ import annotations

import csv
import glob
import json
import math
import os
import re
import zlib
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OPTA_PITCH_LENGTH = 100
OPTA_PITCH_WIDTH = 68
BASE = Path("/Users/user/XG")
UAE_DIR = BASE / "UAE"
MATCHES_CSV = BASE / "UAE Matches.csv"
OUT_DIR = Path("/Users/user/Documents/GitHub/mm-setpieces-1/Data/")


CORNER_HEADERS = [
    "match_id", "Match", "possession", "pass_timestamp", "pass_team_name",
    "Taker", "pass_position", "pass.height.name", "pass.body_part.name",
    "pass.outcome.name", "pass.technique.name", "pass_location_x", "pass_location_y",
    "pass_end_location_x", "pass_end_location_y", "shot_timestamp", "shot_team_name",
    "Shooter", "shot_position", "shot.body_part.name", "shot.outcome.name",
    "shot.statsbomb_xg", "shot_location_x", "shot_location_y", "shot_location_z",
    "Defensive_setup", "Minute", "Second", "SP_outcome",
]

SP_HEADERS = [
    "match_id", "possession", "team.name", "type.name", "SP_Type",
    "location.pass", "pass.height.name", "timestamp", "Taker", "Shooter",
    "location.shot", "shot.statsbomb_xg", "shot.freeze_frame", "shot.outcome.name",
    "shot_x", "shot_y", "Metrics", "Occupation_Rating", "Proximity_Rating",
    "Duel_Win_Prob", "OPS_Opponent_Rating",
]

# Full SP schema required by the repo (superset of SP_HEADERS)
SP_PARQUET_HEADERS = SP_HEADERS + [
    "Restart_Profile", "Start_Third", "Next_3_Box_Entry", "Next_3_Retain_Possession",
    "restart_x", "restart_y", "actions_checked", "delivery_end_x", "delivery_end_y",
]

SET_PIECE_QUALIFIERS = {
    6: "From Corner",
    5: "From Free Kick",
    107: "From Throw In",
}

SHOT_TYPE_TO_SP = {
    "FromCorner":       "From Corner",
    "SetPiece":         "From Free Kick",
    "DirectFreekick":   "From Free Kick",
    "ThrowinSetPiece":  "From Throw In",
}


@dataclass
class MatchInfo:
    match_id: str
    date: str
    match_name: str
    home_name: str
    away_name: str
    home_id: str
    away_id: str


def _opta_id_to_int(opta_id: str) -> int:
    return zlib.crc32(opta_id.encode()) & 0x7FFFFFFF


def norm_name(value: str) -> str:
    value = value.lower()
    value = re.sub(r"\b(fc|sc|club)\b", "", value)
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def parse_filename(path: Path) -> tuple[str, str, str]:
    stem = path.stem
    date, teams = stem.split("_", 1)
    home, away = teams.split(" - ", 1)
    return date, home, away


def to_seconds(minute, second) -> float:
    return float(minute or 0) * 60.0 + float(second or 0)


def timestamp(minute, second) -> str:
    total = int(to_seconds(minute, second))
    ms = int(round((to_seconds(minute, second) - total) * 1000))
    return f"{total // 3600:02d}:{(total % 3600) // 60:02d}:{total % 60:02d}.{ms:03d}"


def opta_x(x) -> float | None:
    if x in ("", None):
        return None
    return round(float(x) * (OPTA_PITCH_LENGTH / 100), 1)


def opta_y(y) -> float | None:
    if y in ("", None):
        return None
    return round(float(y) * (OPTA_PITCH_WIDTH / 100), 1)


def shot_x(x) -> float | None:
    return opta_x(x)


def shot_y(y) -> float | None:
    return opta_y(y)


def fmt_pair(x, y) -> str:
    if x is None or y is None:
        return ""
    return f"{x:g}, {y:g}"


def _start_third(x: float | None) -> str | None:
    if x is None:
        return None
    pct = x / OPTA_PITCH_LENGTH * 100
    if pct < 33.3:
        return "Defensive third"
    if pct < 66.7:
        return "Middle third"
    return "Attacking third"


def read_matches() -> list[MatchInfo]:
    matches = []
    with MATCHES_CSV.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            home = row["matchInfo/contestant/0/name"]
            away = row["matchInfo/contestant/1/name"]
            matches.append(MatchInfo(
                match_id=row["matchInfo/id"],
                date=row["matchInfo/localDate"],
                match_name=f"{home} - {away}",
                home_name=home,
                away_name=away,
                home_id=row["matchInfo/contestant/0/id"],
                away_id=row["matchInfo/contestant/1/id"],
            ))
    return matches


def match_for_file(path: Path, matches: list[MatchInfo]) -> MatchInfo:
    date, file_home, file_away = parse_filename(path)
    home_norm = norm_name(file_home)
    away_norm = norm_name(file_away)
    same_day = [m for m in matches if m.date == date]
    for m in same_day:
        mh = norm_name(m.home_name)
        ma = norm_name(m.away_name)
        if (mh in home_norm or home_norm in mh) and (ma in away_norm or away_norm in ma):
            return MatchInfo(m.match_id, m.date, f"{file_home} - {file_away}", file_home, file_away, m.home_id, m.away_id)
    for m in same_day:
        if norm_name(m.home_name) in home_norm and (norm_name(m.away_name) in away_norm or norm_name(m.away_name).replace("alittihad", "") in away_norm):
            return MatchInfo(m.match_id, m.date, f"{file_home} - {file_away}", file_home, file_away, m.home_id, m.away_id)
    return MatchInfo(path.stem, date, f"{file_home} - {file_away}", file_home, file_away, "", "")


def read_shots() -> dict[str, list[dict]]:
    by_match = defaultdict(list)
    for path in sorted((UAE_DIR / "xgCSV").glob("*.csv")):
        date, home, away = parse_filename(path)
        with path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                sp_type = SHOT_TYPE_TO_SP.get(row.get("Type_of_play", ""))
                if not sp_type:
                    continue
                row["source_file"] = path.name
                row["match_key"] = path.stem
                row["file_date"] = date
                row["file_home"] = home
                row["file_away"] = away
                row["sp_type"] = sp_type
                row["sec"] = to_seconds(row.get("timeMin"), row.get("timeSec"))
                by_match[path.stem].append(row)
    return by_match


def qmap(event: dict) -> dict[int, str | None]:
    return {int(q["qualifierId"]): q.get("value") for q in event.get("qualifier", [])}


def pass_height(quals: dict[int, str | None]) -> str:
    if 2 in quals or 6 in quals:
        return "High Pass"
    if 107 in quals:
        return "Throw-in"
    return "Ground Pass"


def body_part(quals: dict[int, str | None]) -> str:
    if 15 in quals:
        return "Head"
    if 3 in quals:
        return "Right Foot"
    if 152 in quals:
        return "Left Foot"
    return ""


def pass_technique(quals: dict[int, str | None], event: dict) -> str:
    if 6 not in quals:
        return ""
    y = float(event.get("y") or 0)
    end_y = float(quals.get(141) or y)
    if y < 50 and end_y > 40:
        return "Inswinging"
    if y > 50 and end_y < 60:
        return "Inswinging"
    return "Outswinging"


def pass_outcome(event: dict) -> str | None:
    return None if int(event.get("outcome", 0)) == 1 else "Incomplete"


def shot_outcome(shot: dict) -> str:
    if str(shot.get("isGoal", "")).upper() == "TRUE" or str(shot.get("Goal", "")) == "1":
        return "Goal"
    if str(shot.get("isOwnGoal", "")).upper() == "TRUE":
        return "Own Goal"
    return "No Goal"


def read_events(path: Path) -> list[dict]:
    with path.open(encoding="utf-8") as f:
        data = json.load(f)
    events = []
    for event in data.get("event", []):
        event = dict(event)
        event["sec"] = to_seconds(event.get("timeMin"), event.get("timeSec"))
        event["quals"] = qmap(event)
        events.append(event)
    events.sort(key=lambda e: (e["sec"], e.get("eventId", 0), e.get("id", 0)))
    return events


def team_name_for(contestant_id: str, match: MatchInfo) -> str:
    if contestant_id == match.home_id:
        return match.home_name
    if contestant_id == match.away_id:
        return match.away_name
    return contestant_id or ""


def find_matching_shots(start_event: dict, sp_type: str, shots: list[dict], team_name: str, window: int = 30) -> list[dict]:
    start = start_event["sec"]
    return [
        s for s in shots
        if s["sp_type"] == sp_type
        and s.get("TeamId") == team_name
        and start <= s["sec"] <= start + window
    ]


def set_piece_starts(events: list[dict]) -> list[tuple[dict, str]]:
    starts = []
    for event in events:
        if int(event.get("typeId", 0)) != 1:
            continue
        quals = event["quals"]
        for qid, sp_type in SET_PIECE_QUALIFIERS.items():
            if qid in quals:
                starts.append((event, sp_type))
                break
    return starts


# ── Corner rows ───────────────────────────────────────────────────────────────

def build_corner_rows(matches: list[MatchInfo], shots_by_match: dict[str, list[dict]]) -> list[list]:
    rows = []
    for json_path in sorted(UAE_DIR.glob("*.json")):
        match = match_for_file(json_path, matches)
        events = read_events(json_path)
        shots = shots_by_match.get(json_path.stem, [])
        seq = 0
        for event, sp_type in set_piece_starts(events):
            if sp_type != "From Corner":
                continue
            seq += 1
            team = team_name_for(event.get("contestantId"), match)
            matching_shots = find_matching_shots(event, sp_type, shots, team)
            if not matching_shots:
                rows.append(corner_row(match, seq, event, team, None, "No first contact - no shot"))
                continue
            first_delta = min(s["sec"] - event["sec"] for s in matching_shots)
            sp_outcome = "First contact - shot within 3 seconds" if first_delta <= 3.1 else "No first contact - shot"
            for shot in matching_shots:
                rows.append(corner_row(match, seq, event, team, shot, sp_outcome))
    return rows


def corner_row(match: MatchInfo, seq: int, event: dict, team: str, shot: dict | None, sp_outcome: str) -> list:
    quals = event["quals"]
    return [
        match.match_id,
        match.match_name,
        seq,
        timestamp(event.get("timeMin"), event.get("timeSec")),
        team,
        event.get("playerName") or "",
        "",
        pass_height(quals),
        body_part(quals),
        pass_outcome(event),
        pass_technique(quals, event),
        opta_x(event.get("x")),
        opta_y(event.get("y")),
        opta_x(quals.get(140)),
        opta_y(quals.get(141)),
        timestamp(shot.get("timeMin"), shot.get("timeSec")) if shot else None,
        shot.get("TeamId") if shot else None,
        shot.get("PlayerId") if shot else None,
        "",
        shot.get("Bodypart") if shot else None,
        shot_outcome(shot) if shot else None,
        float(shot.get("xG")) if shot and shot.get("xG") else None,
        shot_x(shot.get("x")) if shot else None,
        shot_y(shot.get("y")) if shot else None,
        0 if shot else None,
        "",
        int(float(event.get("timeMin") or 0)),
        float(event.get("timeSec") or 0),
        sp_outcome,
    ]


# ── SP rows ───────────────────────────────────────────────────────────────────

def build_sp_rows(matches: list[MatchInfo], shots_by_match: dict[str, list[dict]]) -> list[list]:
    rows = []
    for json_path in sorted(UAE_DIR.glob("*.json")):
        match = match_for_file(json_path, matches)
        events = read_events(json_path)
        shots = shots_by_match.get(json_path.stem, [])
        seq = 0
        for start, sp_type in set_piece_starts(events):
            seq += 1
            team = team_name_for(start.get("contestantId"), match)
            matching_shots = find_matching_shots(start, sp_type, shots, team)
            if not matching_shots:
                continue
            first_shot = min(matching_shots, key=lambda s: s["sec"])
            sequence_events = [
                e for e in events
                if start["sec"] <= e["sec"] <= first_shot["sec"]
                and e.get("contestantId") == start.get("contestantId")
                and int(e.get("typeId", 0)) == 1
            ]
            if not sequence_events:
                sequence_events = [start]
            for event in sequence_events:
                rows.append(sp_row(match, seq, event, team, sp_type, first_shot))
    return rows


def sp_row(match: MatchInfo, seq: int, event: dict, team: str, sp_type: str, shot: dict) -> list:
    quals = event["quals"]
    px = opta_x(event.get("x"))
    py = opta_y(event.get("y"))
    sx = shot_x(shot.get("x"))
    sy = shot_y(shot.get("y"))
    ex = opta_x(quals.get(140))
    ey = opta_y(quals.get(141))
    # Excel columns (21)
    excel_cols = [
        match.match_id, seq, team, "Pass", sp_type,
        fmt_pair(px, py), pass_height(quals),
        timestamp(event.get("timeMin"), event.get("timeSec")),
        event.get("playerName") or "", shot.get("PlayerId") or "",
        fmt_pair(sx, sy),
        float(shot.get("xG")) if shot.get("xG") else None,
        "", shot_outcome(shot), sx, sy,
        "", "", "", "", "",
    ]
    # Extra parquet-only columns (9)
    parquet_extra = [
        None,              # Restart_Profile
        _start_third(px), # Start_Third
        False,             # Next_3_Box_Entry
        False,             # Next_3_Retain_Possession
        px,                # restart_x
        py,                # restart_y
        0,                 # actions_checked
        ex,                # delivery_end_x
        ey,                # delivery_end_y
    ]
    return excel_cols + parquet_extra


# ── Parquet writers ───────────────────────────────────────────────────────────

def _to_int64(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").astype("Int64")


def write_corners_parquet(rows: list[list], path: Path) -> None:
    df = pd.DataFrame(rows, columns=CORNER_HEADERS)
    # Convert Opta alphanumeric match_id → stable int
    df["match_id"] = df["match_id"].apply(
        lambda v: _opta_id_to_int(str(v)) if isinstance(v, str) and not str(v).isdigit() else int(v)
    ).astype("int64")
    df["possession"] = _to_int64(df["possession"])
    df["Minute"]     = _to_int64(df["Minute"])
    df["Second"]     = pd.to_numeric(df["Second"], errors="coerce")
    for col in ("pass_location_x", "pass_location_y", "pass_end_location_x",
                "pass_end_location_y", "shot_location_x", "shot_location_y",
                "shot_location_z", "shot.statsbomb_xg"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(path, engine="pyarrow", compression="zstd", index=False)
    print(f"  {path.name}: {len(df)} rows")


def write_sp_parquet(rows: list[list], fk_path: Path, ti_path: Path) -> None:
    df = pd.DataFrame(rows, columns=SP_PARQUET_HEADERS)
    df["match_id"] = df["match_id"].apply(
        lambda v: _opta_id_to_int(str(v)) if isinstance(v, str) and not str(v).isdigit() else int(v)
    ).astype("int64")
    df["possession"]          = _to_int64(df["possession"])
    df["Occupation_Rating"]   = pd.to_numeric(df["Occupation_Rating"],   errors="coerce").fillna(0).astype("int64")
    df["Duel_Win_Prob"]       = pd.to_numeric(df["Duel_Win_Prob"],       errors="coerce").fillna(0).astype("int64")
    df["OPS_Opponent_Rating"] = pd.to_numeric(df["OPS_Opponent_Rating"], errors="coerce").fillna(0).astype("int64")
    df["actions_checked"]     = pd.to_numeric(df["actions_checked"],     errors="coerce").fillna(0).astype("int64")
    df["Next_3_Box_Entry"]         = df["Next_3_Box_Entry"].astype(bool)
    df["Next_3_Retain_Possession"] = df["Next_3_Retain_Possession"].astype(bool)
    for col in ("shot.statsbomb_xg", "shot_x", "shot_y",
                "restart_x", "restart_y", "delivery_end_x", "delivery_end_y"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    fk = df[df["SP_Type"] == "From Free Kick"].reset_index(drop=True)
    ti = df[df["SP_Type"] == "From Throw In"].reset_index(drop=True)

    for path, subset in ((fk_path, fk), (ti_path, ti)):
        path.parent.mkdir(parents=True, exist_ok=True)
        subset.to_parquet(path, engine="pyarrow", compression="zstd", index=False)
        print(f"  {path.name}: {len(subset)} rows")


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_workbook(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.append(headers)
    for row in rows:
        ws.append(row[:len(headers)])  # Excel only gets its column subset

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 48)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float) and not math.isclose(cell.value, round(cell.value)):
                cell.number_format = "0.000"
    ws.sheet_view.showGridLines = True
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    matches = read_matches()
    shots_by_match = read_shots()

    print("Building corner rows …")
    corner_rows = build_corner_rows(matches, shots_by_match)
    print(f"  {len(corner_rows)} rows")

    print("Building SP rows …")
    sp_rows = build_sp_rows(matches, shots_by_match)
    print(f"  {len(sp_rows)} rows")

    # Excel output (unchanged)
    corner_xlsx = OUT_DIR / "Corners/UAE - Corners 2025-2026.xlsx"
    sp_xlsx     = OUT_DIR / "UAE SP.xlsx"
    write_workbook(corner_xlsx, CORNER_HEADERS, corner_rows)
    write_workbook(sp_xlsx,     SP_HEADERS,     sp_rows)

    for output in (corner_xlsx, sp_xlsx):
        wb = load_workbook(output, read_only=True, data_only=True)
        ws = wb.active
        print(f"{output.name}: {ws.max_row - 1} data rows, {ws.max_column} columns")
        wb.close()

    # Parquet output
    print("\nWriting parquet files …")
    write_corners_parquet(
        corner_rows,
        OUT_DIR / "Corners/UAE - Corners 2025-2026.parquet",
    )
    write_sp_parquet(
        sp_rows,
        fk_path=OUT_DIR / "SP/UAE - Freekicks.parquet",
        ti_path=OUT_DIR / "SP/UAE - Throwins.parquet",
    )
    print("Done.")


if __name__ == "__main__":
    main()
